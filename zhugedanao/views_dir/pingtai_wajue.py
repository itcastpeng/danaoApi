from openpyxl.styles import Font, Alignment
from openpyxl import Workbook
import os, time
from zhugedanao import models
from publicFunc import Response
from publicFunc import account
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt, csrf_protect
import datetime
from zhugedanao.forms.pingtai_wajue import AddForm, SelectForm
import json
import random
from django.db.models import Count,Q, Sum
import threading
import redis

redis_rc = redis.Redis(host='redis_host', port=6379, db=3, decode_responses=True)
# redis_rc = redis.Redis(host='192.168.100.20', port=6379, db=4, decode_responses=True)

# cerf  token验证 用户展示模块
@csrf_exempt
@account.is_token(models.zhugedanao_userprofile)
def pingTaiWaJueShow(request):
    response = Response.ResponseObj()
    user_id = request.GET.get('user_id')
    if request.method == "GET":
        forms_obj = SelectForm(request.GET)
        if forms_obj.is_valid():
            current_page = forms_obj.cleaned_data['current_page']
            length = forms_obj.cleaned_data['length']
            task_objs = models.zhugedanao_pingtaiwajue_keyword.objects.filter(user_id_id=user_id)
            yuming_objs = models.zhugedanao_pingtaiwajue_yuming.objects.filter(tid__user_id=user_id)
            yuming_count = yuming_objs.count()

            task_count = task_objs.count()
            yiwancheng = task_objs.filter(is_perform=1).count()
            query_progress = 0
            if yiwancheng:
                query_progress = int((int(yiwancheng) / int(task_count)) * 100)
            whether_complete = False
            if yiwancheng == task_count:
                whether_complete = True
            pingtaiwajue_chongfu = redis_rc.get('danao_pingtaiwajue_chongfu')
            chongfu = 0
            if pingtaiwajue_chongfu:
                chongfu = pingtaiwajue_chongfu
            # 分页
            if length != 0:
                start_line = (current_page - 1) * length
                stop_line = start_line + length
                objs = yuming_objs[start_line: stop_line]
            data_list = []

            for obj in objs:
                if str(obj.tid.search) == '1':
                    yinqing = '百度'
                elif str(obj.tid.search) == '4':
                    yinqing = '手机百度'
                elif str(obj.tid.search) == '3':
                    yinqing = '360'
                elif str(obj.tid.search) == '6':
                    yinqing = '手机360'
                else:
                    yinqing = ''
                data_list.append({
                    'yuming':obj.yuming,
                    'number':obj.number,
                    'search':yinqing
                })
            response.code = 200
            response.msg = '查询成功'
            response.data = {
                'yuming_count':yuming_count,            # 域名数量
                'data': data_list,                      # 详情
                'objs_count':task_count,                # 总数
                'query_progress':query_progress,        # 进度
                'whether_complete':whether_complete,    # 是否完成
                'yiwancheng_obj':yiwancheng,            # 已完成
                'chongfu_num':int(chongfu)
            }
        else:
            response.code = 402
            response.msg = "请求异常"
            response.data = json.loads(forms_obj.errors.as_json())
    return JsonResponse(response.__dict__)



#  增删改
#  csrf  token验证
@csrf_exempt
@account.is_token(models.zhugedanao_userprofile)
def pingTaiWaJue(request, oper_type, o_id):
    response = Response.ResponseObj()
    user_id = request.GET.get('user_id')
    if request.method == "POST":
        # 增加下拉任务
        if oper_type == "add":
            models.zhugedanao_pingtaiwajue_yuming.objects.filter(tid__user_id_id=user_id).delete()
            models.zhugedanao_pingtaiwajue_keyword.objects.filter(user_id_id=user_id).delete()
            form_data = {
                'search' : request.POST.get('search'),
                'keywords': request.POST.get('keywords'),
                'page_number': request.POST.get('page_number'),
            }
            #  创建 form验证 实例（参数默认转成字典）
            forms_obj = AddForm(form_data)
            if forms_obj.is_valid():
                #  添加数据库
                chongfu = int(len(forms_obj.cleaned_data.get('keywords'))) - int(len(set(forms_obj.cleaned_data.get('keywords'))))
                print('chongfu============>',chongfu)
                redis_rc.set('danao_pingtaiwajue_chongfu', '{}'.format(int(chongfu)), ex=None, px=None, nx=False, xx=False)
                keywords_list = set(forms_obj.cleaned_data.get('keywords'))
                querysetlist = []
                create_time = datetime.datetime.today().strftime('%Y-%m-%d %H:%M:%S')
                for search in json.loads(form_data['search']):
                    for keyword in keywords_list:
                        querysetlist.append(
                            models.zhugedanao_pingtaiwajue_keyword(
                                user_id_id=user_id,
                                search=search,
                                keyword=keyword,
                                create_time=create_time,
                                page_number=json.loads(form_data['page_number'])
                            )
                        )
                models.zhugedanao_pingtaiwajue_keyword.objects.bulk_create(querysetlist)
                response.code = 200
                response.msg = "添加成功"
                response.data = {}
            else:
                print("验证不通过")
                response.code = 301
                response.msg = json.loads(forms_obj.errors.as_json())
                response.data = {}

    elif request.method == 'GET':
        # 点击返回 删除任务
        if oper_type == 'clickReturn':
            response.code = 200
            response.msg = '退出成功'
            models.zhugedanao_pingtaiwajue_yuming.objects.filter(
                tid__user_id_id=user_id
            ).delete()
            models.zhugedanao_pingtaiwajue_keyword.objects.filter(
                user_id_id=user_id
            ).delete()
            return JsonResponse(response.__dict__)

        # 生成excel
        if oper_type == 'generateExcel':
            now_date = datetime.datetime.today().strftime('%Y-%m-%d %H-%M-%S')
            wb = Workbook()
            ws = wb.active
            ws.cell(row=1, column=1, value="任务名称")
            ws.cell(row=3, column=1, value="序号")
            ws.cell(row=3, column=2, value="平台名称")
            ws.cell(row=3, column=3, value="排名数")
            ws.cell(row=3, column=4, value="比例")
            ws.cell(row=3, column=5, value="搜索引擎")
            ws.cell(row=3, column=6, value="查询时间:{}".format(now_date))

            # 合并单元格        开始行      结束行       用哪列          占用哪列
            ws.merge_cells(start_row=1, end_row=2, start_column=1, end_column=8)
            ws.merge_cells(start_row=3, end_row=3, start_column=1, end_column=1)
            ws.merge_cells(start_row=3, end_row=3, start_column=2, end_column=2)
            ws.merge_cells(start_row=3, end_row=3, start_column=3, end_column=3)
            ws.merge_cells(start_row=3, end_row=3, start_column=4, end_column=4)
            ws.merge_cells(start_row=3, end_row=3, start_column=5, end_column=5)
            ws.merge_cells(start_row=3, end_row=3, start_column=6, end_column=8)
            ws.merge_cells(start_row=4, end_row=4, start_column=6, end_column=8)
            ws.merge_cells(start_row=5, end_row=5, start_column=6, end_column=8)

            # print('设置列宽')
            ws.column_dimensions['A'].width = 9
            ws.column_dimensions['B'].width = 30
            ws.column_dimensions['C'].width = 9
            ws.column_dimensions['D'].width = 20
            ws.column_dimensions['E'].width = 9
            ws.column_dimensions['F'].width = 10
            ws.column_dimensions['H'].width = 70

            # print('设置行高')
            # ws.row_dimensions[1].height = 28

            # print('文本居中')
            ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
            ws['D2'].alignment = Alignment(horizontal='center', vertical='center')
            ws['E2'].alignment = Alignment(horizontal='center', vertical='center')
            ws['F3'].alignment = Alignment(horizontal='center', vertical='center')
            ws['F4'].alignment = Alignment(horizontal='center', vertical='center')
            ws['F5'].alignment = Alignment(horizontal='center', vertical='center')
            row = 4
            objs = models.zhugedanao_pingtaiwajue_yuming.objects.filter(tid__user_id_id=user_id).order_by('-number')
            # print('objs.count()=========> ', objs.count())
            number_count = objs.values('tid__user_id').annotate(Sum('number'))
            # print('number_count==========> ',number_count)
            number = 0
            bili = 0
            yinqing = '百度'
            for obj in objs:
                if str(obj.tid.search) == '1':
                    yinqing = '百度'
                elif str(obj.tid.search) == '4':
                    yinqing = '手机百度'
                elif str(obj.tid.search) == '3':
                    yinqing = '360'
                elif str(obj.tid.search) == '6':
                    yinqing = '手机360'
                number += 1
                num_count = int(number_count[0]['number__sum'])
                if obj.number:
                    bili = round(int(obj.number) / num_count, 2)
                print('bilibili===============> ',bili)
                ws.cell(row=row, column=1, value="{number}".format(number=number))
                ws.cell(row=row, column=2, value="{title}".format(title=obj.yuming))
                ws.cell(row=row, column=3, value="{title}".format(title=obj.number))
                ws.cell(row=row, column=4, value="{bili}".format(bili=str(bili) + '%'))
                ws.cell(row=row, column=5, value="{search}".format(search=yinqing))
                row += 1
            ws.cell(row=4, column=6, value="总排名数:{}".format(number_count[0]['number__sum']))
            ws.cell(row=5, column=6, value="平台数:{}".format(objs.count()))
            randInt = random.randint(1, 100)
            nowDateTime = int(time.time())
            excel_name = str(randInt) + str(nowDateTime)
            wb.save(os.path.join(os.getcwd(), 'statics', 'zhugedanao', 'pingTaiWaJueExcel' , '{}.xlsx'.format(excel_name)))
            response.code = 200
            response.msg = '生成成功'
            response.data = {'excel_name':'http://api.zhugeyingxiao.com/' + os.path.join('statics', 'zhugedanao', 'pingTaiWaJueExcel' , '{}.xlsx'.format(excel_name))}
            return JsonResponse(response.__dict__)

        # 处理后最终结果
        if oper_type == 'finalResult':
            # while True:
            # keyword_objs = models.zhugedanao_pingtaiwajue_keyword.objects.filter(user_id_id=user_id)
            # yiwancheng = keyword_objs.filter(is_perform=1).count()
            # keyword_all = keyword_objs.count()
            # print('---------> ', yiwancheng, keyword_all)
            # if int(yiwancheng) == int(keyword_all):
            #     break
            # else:
            # objs = models.zhugedanao_pingtaiwajue_yuming.objects.filter(tid__user_id_id=user_id).values('yuming','tid__search').annotate(
            #     Sum('number')
            # )
                # final_objs = models.zhugedanao_pingtaiwajue_finalResult.objects
                # for obj in objs:
                #     is_yuming = final_objs.filter(yuming=obj['yuming'])
                #     print('is_yuming============> ',is_yuming)
                #     if is_yuming:
                #         final_objs.update(
                #             user_id_id=user_id,
                #             yuming=obj['yuming'],
                #             number=obj['number__sum'],
                #             search=obj['tid__search']
                #         )
                #     else:
                #         final_objs.create(
                #             user_id_id=user_id,
                #             yuming=obj['yuming'],
                #             number=obj['number__sum'],
                #             search=obj['tid__search']
                #         )
            objs = models.zhugedanao_pingtaiwajue_yuming.objects
            tongji_objs = objs.filter(tid__user_id_id=user_id).values('yuming',
                'tid__search').annotate(
                Sum('number')
            )
            for tongji_obj in tongji_objs:
                yuming_obj = objs.filter(tid__user_id_id=user_id).filter(yuming=tongji_obj['yuming']).filter(tid__search=tongji_obj['tid__search'])
                if yuming_obj[0].yuming and yuming_obj[0].number != tongji_obj['number__sum']:
                    now_date = datetime.datetime.today().strftime('%Y-%m-%d %H:%M:%S')
                    create_obj = objs.create(
                        yuming=tongji_obj['yuming'],
                        number=tongji_obj['number__sum'],
                        tid_id=yuming_obj[0].tid.id,
                        create_time=now_date
                    )
                    objs.filter(yuming=tongji_obj['yuming']).exclude(id=create_obj.id).delete()
            response.data = {}
            response.code = 2000
            response.msg = '计算结果完成'
    else:
        response.code = 402
        response.msg = "请求异常"



    return JsonResponse(response.__dict__)