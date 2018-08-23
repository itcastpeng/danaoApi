from openpyxl.styles import Font, Alignment
from openpyxl import Workbook
import os, time
from zhugedanao import models
from publicFunc import Response
from publicFunc import account
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt, csrf_protect
import datetime
from zhugedanao.forms.shoulu_chaxun import AddForm, SelectForm
import json
import random

chongfu = 0

# cerf  token验证 用户展示模块
@csrf_exempt
@account.is_token(models.zhugedanao_userprofile)
def shouLuChaXunShow(request):
    response = Response.ResponseObj()
    user_id = request.GET.get('user_id')
    difference_status = request.GET.get('difference_status')
    if request.method == "GET":
        forms_obj = SelectForm(request.GET)
        if forms_obj.is_valid():
            current_page = forms_obj.cleaned_data['current_page']
            length = forms_obj.cleaned_data['length']
            objs = models.zhugedanao_shoulu_chaxun.objects.filter(user_id_id=user_id)
            if difference_status:
                if int(difference_status) == 0:         # 已收录
                    objs = models.zhugedanao_shoulu_chaxun.objects.filter(user_id_id=user_id, is_shoulu=1)
                if int(difference_status) == 1:         # 未收录
                    objs = models.zhugedanao_shoulu_chaxun.objects.filter(user_id_id=user_id, is_shoulu=2)

            dataCount = objs.count()        # 收录数据总数
            # 已完成 进度条
            zhixingCount = objs.filter(is_zhixing=1)
            yiZhiXingCount = zhixingCount.count()
            shouluCount = zhixingCount.filter(is_shoulu=1).count()
            query_progress = 0
            if yiZhiXingCount:
                query_progress = int((yiZhiXingCount / dataCount) * 100)

            # 收录率
            shouLuLv = 0
            if shouluCount:
                shouLuLv = int((shouluCount / dataCount) * 100)

            # 分页
            if length != 0:
                start_line = (current_page - 1) * length
                stop_line = start_line + length
                objs = objs[start_line: stop_line]
            whether_complete = False
            if dataCount == yiZhiXingCount:
                whether_complete = True
            # 返回的数据
            retData = []
            for obj in objs:
                is_shoulu = ''
                if int(obj.is_shoulu) == 1:
                    is_shoulu = True
                if int(obj.is_shoulu) == 2:
                    is_shoulu = False
                if str(obj.search) == '1':
                    yinqing = '百度'
                elif str(obj.search) == '4':
                    yinqing = '手机百度'
                elif str(obj.search) == '3':
                    yinqing = '360'
                elif str(obj.search) == '6':
                    yinqing = '手机360'
                else:
                    yinqing = ''
                retData.append({
                    'shoulu_status':is_shoulu,
                    'website':obj.url,
                    'title':obj.title,
                    'search_engine':yinqing,
                    'kuaizhao_date':obj.kuaizhao_time,
                    'statusCode':obj.status_code,
                })
            response.code = 200
            response.msg = '查询成功'
            response.data = {
                'data': retData,                        # 详情
                'count_obj': dataCount,                 # 任务总数
                'shoulushu':shouluCount,                # 收录数
                'shoululv':shouLuLv,                    # 收录率
                'yiwancheng_obj':yiZhiXingCount,        # 已完成数量
                'query_progress':query_progress,        # 进度条
                'whether_complete':whether_complete,    # 是否全部完成
                'chongfu_num':chongfu,                        # 重复数
            }
        else:
            response.code = 402
            response.msg = "请求异常"
            response.data = json.loads(forms_obj.errors.as_json())

    return JsonResponse(response.__dict__)


#  增删改
#  csrf  token验证
@csrf_exempt
@account.is_token(models.zhugedanao_userprofile)
def shouLuChaxun(request, oper_type, o_id):
    response = Response.ResponseObj()
    user_id = request.GET.get('user_id')
    if request.method == "POST":
        # 增加收录任务
        if oper_type == "add":
            models.zhugedanao_shoulu_chaxun.objects.filter(user_id_id=user_id).delete()
            form_data = {
                'search_list': request.POST.get('search_list'),
                'url_list': request.POST.get('url_list'),
            }
            #  创建 form验证 实例（参数默认转成字典）
            forms_obj = AddForm(form_data)
            if forms_obj.is_valid():
                global chongfu
                print("验证通过")
                #  添加数据库
                print('-------->', forms_obj.cleaned_data.get('url_list'), type(forms_obj.cleaned_data.get('url_list')))
                chongfu = int(len(forms_obj.cleaned_data.get('url_list'))) - int(len(set(forms_obj.cleaned_data.get('url_list'))))
                url_list = set(forms_obj.cleaned_data.get('url_list'))
                search_list = forms_obj.cleaned_data.get('search_list')
                querysetlist = []
                now_date = datetime.datetime.today().strftime('%Y-%m-%d %H:%M:%S')
                for search in json.loads(search_list):
                    for url in url_list:
                        querysetlist.append(
                            models.zhugedanao_shoulu_chaxun(
                                user_id_id=user_id,
                                url=url,
                                search=search,
                                createAndStart_time=now_date,
                            )
                        )
                models.zhugedanao_shoulu_chaxun.objects.bulk_create(querysetlist)
                response.code = 200
                response.msg = "添加成功"
                response.data = {}
            else:
                print("验证不通过")
                response.code = 301
                response.msg = json.loads(forms_obj.errors.as_json())
                response.data = {}

    elif request.method == 'GET':

        # 点击返回 删除任务
        if oper_type == 'clickReturn':
            response.code = 200
            response.msg = '退出成功'
            print('user_id=====> ',user_id)
            models.zhugedanao_shoulu_chaxun.objects.filter(user_id_id=user_id).delete()
            return JsonResponse(response.__dict__)

            # 生成报表

        # 生成excel
        if oper_type == 'generateExcel':
            now_date = datetime.datetime.today().strftime('%Y-%m-%d %H-%M-%S')
            wb = Workbook()
            ws = wb.active
            ws.cell(row=1, column=1, value="收录查询")
            ws.cell(row=2, column=4, value="查询时间:")
            ws.cell(row=6, column=1, value="标题")
            ws.cell(row=6, column=2, value="网址")
            ws.cell(row=6, column=3, value="搜索引擎")
            ws.cell(row=6, column=4, value="收录")
            ws.cell(row=6, column=5, value="快照日期")
            ft1 = Font(name='宋体', size=22)
            a1 = ws['A1']
            a1.font = ft1

            # 合并单元格        开始行      结束行       用哪列          占用哪列
            ws.merge_cells(start_row=1, end_row=1, start_column=1, end_column=6)
            ws.merge_cells(start_row=2, end_row=5, start_column=1, end_column=1)
            ws.merge_cells(start_row=2, end_row=5, start_column=2, end_column=2)
            ws.merge_cells(start_row=2, end_row=5, start_column=3, end_column=3)
            ws.merge_cells(start_row=2, end_row=5, start_column=4, end_column=4)
            ws.merge_cells(start_row=2, end_row=5, start_column=5, end_column=5)

            # print('设置列宽')
            ws.column_dimensions['A'].width = 30
            ws.column_dimensions['B'].width = 30
            ws.column_dimensions['C'].width = 30
            ws.column_dimensions['D'].width = 30
            ws.column_dimensions['E'].width = 30
            ws.column_dimensions['F'].width = 30

            # print('设置行高')
            ws.row_dimensions[1].height = 28

            # print('文本居中')
            ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
            ws['D2'].alignment = Alignment(horizontal='center', vertical='center')
            ws['E2'].alignment = Alignment(horizontal='center', vertical='center')
            row = 7
            objs = models.zhugedanao_shoulu_chaxun.objects.filter(user_id_id=user_id)
            for obj in objs:
                is_shoulu = '未收录'
                if str(obj.is_shoulu) == '1':
                    is_shoulu = '已收录'
                if str(obj.search) == '1':
                    yinqing = '百度'
                elif str(obj.search) == '4':
                    yinqing = '手机百度'
                elif str(obj.search) == '3':
                    yinqing = '360'
                elif str(obj.search) == '6':
                    yinqing = '手机360'
                else:
                    yinqing = ''
                ws.cell(row=2, column=5, value="{chaxun_time}".format(chaxun_time=obj.createAndStart_time))
                ws.cell(row=row, column=1, value="{title}".format(title=obj.title))
                ws.cell(row=row, column=2, value="{url}".format(url=obj.url))
                ws.cell(row=row, column=3, value="{search}".format(search=yinqing))
                ws.cell(row=row, column=4, value="{is_shoulu}".format(is_shoulu=is_shoulu))
                ws.cell(row=row, column=5, value="{kuaizhao_time}".format(kuaizhao_time=obj.kuaizhao_time))
                row += 1
            randInt = random.randint(1, 100)
            nowDateTime = int(time.time())
            excel_name = str(randInt) + str(nowDateTime)
            wb.save(os.path.join(os.getcwd(), 'statics', 'zhugedanao', 'shouLuExcel' , '{}.xlsx'.format(excel_name)))
            # print('==========>','http://api.zhugeyingxiao.com/' + os.path.join('statics', 'zhugedanao', 'shouLuExcel' , '{}.xlsx'.format(excel_name)))
            # print('==========>','http://127.0.0.1:8000/' + os.path.join('statics', 'zhugedanao', 'shouLuExcel' , '{}.xlsx'.format(excel_name)))
            response.code = 200
            response.msg = '生成成功'
            response.data = {'excel_name':'http://api.zhugeyingxiao.com/' + os.path.join('statics', 'zhugedanao', 'shouLuExcel' , '{}.xlsx'.format(excel_name))}
            return JsonResponse(response.__dict__)

    else:
        response.code = 402
        response.msg = "请求异常"


    return JsonResponse(response.__dict__)