from openpyxl.styles import Font, Alignment
from openpyxl import Workbook
import os, time
from zhugedanao import models
from publicFunc import Response
from publicFunc import account
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt, csrf_protect
import datetime
from zhugedanao.forms.baidu_xiala import AddForm, SelectForm
import json
import random
from django.db.models import Count,Q, Sum
import threading
import redis

redis_rc = redis.Redis(host='redis://redis_host', port=6379, db=4, decode_responses=True)
# redis_rc = redis.Redis(host='192.168.100.20', port=6379, db=4, decode_responses=True)
# cerf  token验证 用户展示模块
@csrf_exempt
@account.is_token(models.zhugedanao_userprofile)
def baiDuXiaLaShow(request):
    response = Response.ResponseObj()
    user_id = request.GET.get('user_id')
    if request.method == "GET":
        forms_obj = SelectForm(request.GET)
        if forms_obj.is_valid():
            current_page = forms_obj.cleaned_data['current_page']
            length = forms_obj.cleaned_data['length']
            objs = models.zhugedanao_baiduxiala_chaxun.objects.filter(user_id_id=user_id)
            obj_count = objs.count()
            zhixing_count = objs.filter(is_zhixing=1).count()
            num = 0
            whether_complete = False
            if int(zhixing_count) == int(obj_count):
                whether_complete = True
            query_progress = 0
            if zhixing_count:
                query_progress = int((int(zhixing_count) / int(obj_count)) * 100)
            for obj in objs:
                if obj.xialaci:
                    for xiala in eval(obj.xialaci):
                        num += 1
            baiduxiala_chongfu = redis_rc.get('danao_baiduxiala_chongfu')
            chongfu = 0
            if baiduxiala_chongfu:
                chongfu = baiduxiala_chongfu

            # 分页
            if length != 0:
                start_line = (current_page - 1) * length
                stop_line = start_line + length
                objs = objs[start_line: stop_line]
            data_list = []
            for obj in objs:
                xialaci = ''
                if obj.xialaci:
                    xialaci = eval(obj.xialaci)
                data_list.append({
                    'keyword':obj.keyword,
                    'search': obj.search,
                    'otherData': xialaci
                })
            response.code = 200
            response.msg = '查询成功'
            response.data = {
                'retData': data_list,
                'keyword_count':obj_count,
                'obj_count':num,
                'yiwancheng_obj':zhixing_count,
                'whether_complete': whether_complete,  # 是否全部完成
                'query_progress': query_progress,  # 进度条
                'chongfu_num':int(chongfu)
            }
        else:
            response.code = 402
            response.msg = "请求异常"
            response.data = json.loads(forms_obj.errors.as_json())
    return JsonResponse(response.__dict__)


#  增删改
#  csrf  token验证
@csrf_exempt
@account.is_token(models.zhugedanao_userprofile)
def baiDuXiaLa(request, oper_type, o_id):
    response = Response.ResponseObj()
    user_id = request.GET.get('user_id')
    if request.method == "POST":
        # 增加下拉任务
        if oper_type == "add":
            form_data = {
                'search_list': request.POST.get('searchEngineModel'),
                'keywords_list': request.POST.get('editor_content'),
            }
            querysetlist = []
            forms_obj = AddForm(form_data)
            if forms_obj.is_valid():
                models.zhugedanao_baiduxiala_chaxun.objects.filter(user_id_id=user_id).delete()
                chongfu = int(len(forms_obj.cleaned_data.get('keywords_list'))) - int(len(set(forms_obj.cleaned_data.get('keywords_list'))))
                print('chongfu=======> ',chongfu)
                redis_rc.set('danao_baiduxiala_chongfu', '{}'.format(chongfu), ex=None, px=None, nx=False, xx=False)
                now_date = datetime.datetime.today().strftime('%Y-%m-%d %H:%M:%S')
                for search in eval(forms_obj.cleaned_data.get('search_list')):
                    for keyword in set(forms_obj.cleaned_data.get('keywords_list')):
                        querysetlist.append(
                            models.zhugedanao_baiduxiala_chaxun(
                                user_id_id = user_id,
                                keyword = keyword,
                                search = search,
                                is_zhixing = 0,
                                createAndStart_time=now_date
                            )
                        )
                models.zhugedanao_baiduxiala_chaxun.objects.bulk_create(querysetlist)
                response.code = 200
                response.msg = "添加成功"
                response.data = {}
            else:
                print("验证不通过")
                response.code = 301
                response.msg = json.loads(forms_obj.errors.as_json())

    elif request.method == 'GET':
        # 点击返回 删除任务
        if oper_type == 'clickReturn':
            response.code = 200
            response.msg = '退出成功'
            models.zhugedanao_baiduxiala_chaxun.objects.filter(
                user_id_id=user_id
            ).delete()
            return JsonResponse(response.__dict__)

        # 生成excel
        if oper_type == 'generateExcel':
            now_date = datetime.date.today().strftime('%Y-%m-%d')
            wb = Workbook()
            ws = wb.active
            ws.cell(row=1, column=1, value="序号")
            ws.cell(row=1, column=2, value="关键词")
            ws.cell(row=1, column=3, value="下拉词")
            ws.cell(row=1, column=4, value="搜索引擎")
            ws.cell(row=1, column=5, value="制表日期:{}".format(now_date))
            ft1 = Font(name='宋体', size=22)
            # a1 = ws['A1']
            # a1.font = ft1

            # 合并单元格        开始行      结束行       用哪列          占用哪列
            # ws.merge_cells(start_row=1, end_row=2, start_column=1, end_column=8)

            # print('设置列宽')
            ws.column_dimensions['A'].width = 9
            ws.column_dimensions['B'].width = 10
            ws.column_dimensions['C'].width = 20
            ws.column_dimensions['D'].width = 10
            ws.column_dimensions['E'].width = 20

            # print('设置行高')
            ws.row_dimensions[1].height = 20

            # print('文本居中')
            ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
            ws['B1'].alignment = Alignment(horizontal='center', vertical='center')
            ws['C1'].alignment = Alignment(horizontal='center', vertical='center')
            ws['D1'].alignment = Alignment(horizontal='center', vertical='center')
            ws['E1'].alignment = Alignment(horizontal='center', vertical='center')
            row = 2
            objs = models.zhugedanao_baiduxiala_chaxun.objects.filter(user_id_id=user_id)
            yinqing = '百度'
            number = 0
            num = 0
            for obj in objs:
                if int(obj.search) == 1:
                    yinqing = '百度'
                elif int(obj.search) == 4:
                    yinqing = '手机百度'
                number += 1
                number_ws = ws.cell(row=row, column=1, value="{number}".format(number=number))
                number_ws.alignment = Alignment(horizontal='center', vertical='center')
                number_ws.font = Font(size=10)
                if obj.xialaci:
                    for xialaci in eval(obj.xialaci):
                        num += 1
                        ws.cell(row=row, column=2, value="{title}".format(title=obj.keyword)).font = Font(size=10)
                        xialaci_ws = ws.cell(row=row, column=3, value="{title}".format(title=xialaci))
                        xialaci_ws.alignment = Alignment(horizontal='center', vertical='center')
                        xialaci_ws.font = Font(size=10)
                else:
                    ws.cell(row=row, column=2, value="{title}".format(title=obj.keyword)).font = Font(size=10)
                search_ws = ws.cell(row=row, column=4, value="{search}".format(search=yinqing))
                search_ws.alignment = Alignment(horizontal='center', vertical='center')
                search_ws.font = Font(size=10)
                row += 1
            ws.cell(row=2, column=5, value="数据总数：{}".format(objs.count())).font = Font('宋体', color='0066CD', size=10)
            ws.cell(row=3, column=5, value="重复：{}".format(0)).font = Font('宋体', color='0066CD', size=10)
            ws.cell(row=4, column=5, value="下拉数：{}".format(num)).font = Font('宋体', color='0066CD', size=10)
            ws.cell(row=5, column=5, value="异常：{}".format(0)).font = Font('宋体', color='0066CD', size=10)
            ws.cell(row=6, column=5, value="百度下拉数：{}".format(0)).font = Font('宋体', color='0066CD', size=10)

            randInt = random.randint(1, 100)
            nowDateTime = int(time.time())
            excel_name = str(randInt) + str(nowDateTime)
            wb.save(os.path.join(os.getcwd(), 'statics', 'zhugedanao', 'baiDuXiaLa' , '{}.xlsx'.format(excel_name)))
            response.code = 200
            response.msg = '生成成功'
            response.data = {'excel_name':'http://api.zhugeyingxiao.com/' + os.path.join('statics', 'zhugedanao', 'baiDuXiaLa' , '{}.xlsx'.format(excel_name))}
            return JsonResponse(response.__dict__)
    else:
        response.code = 402
        response.msg = "请求异常"

    return JsonResponse(response.__dict__)