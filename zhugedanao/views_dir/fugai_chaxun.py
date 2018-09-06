from openpyxl.styles import Font, Alignment
from openpyxl import Workbook
import os, time
from zhugedanao import models
from publicFunc import Response
from publicFunc import account
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt, csrf_protect
import datetime
from zhugedanao.forms.fugai_chaxun import AddForm, SelectForm
import json
import random

import redis

redis_rc = redis.Redis(host='redis_host', port=6379, db=3, decode_responses=True)
# redis_rc = redis.Redis(host='192.168.100.20', port=6379, db=4, decode_responses=True)

# cerf  token验证 用户展示模块
@csrf_exempt
@account.is_token(models.zhugedanao_userprofile)
def fuGaiChaxunShow(request):
    response = Response.ResponseObj()
    user_id = request.GET.get('user_id')
    if request.method == "GET":
        forms_obj = SelectForm(request.GET)
        if forms_obj.is_valid():
            current_page = forms_obj.cleaned_data['current_page']
            length = forms_obj.cleaned_data['length']
            objs = models.zhugedanao_fugai_chaxun.objects.filter(user_id_id=user_id)
            dataCount = objs.count()        # 覆盖总数
            objs_eidt = objs.filter(is_zhixing=1)
            # 已完成 进度条
            yiZhiXingCount = objs_eidt.count()
            yiWanCheng = 0
            if yiZhiXingCount:
                yiWanCheng = yiZhiXingCount
            whether_complete = False
            if dataCount == yiWanCheng:
                whether_complete = True
            fugai_objs = objs_eidt.filter(paiming_detail__isnull=False).filter(paiming_detail__gt=0)
            paiming_num = fugai_objs.count()
            fugai_num = 0
            for fugai_obj in fugai_objs:
                fugai_num += len(fugai_obj.paiming_detail.split(','))
            fugailv = 0
            if fugai_num:
                fugailv = int((int(fugai_num) / int(dataCount * 10)) * 100)
            paiminglv = 0
            if paiming_num:
                paiminglv = int((int(paiming_num) / int(dataCount)) * 100)
            query_progress = 0 # 进度条
            if yiZhiXingCount:
                query_progress = int((yiZhiXingCount / dataCount) * 100)
            fugai_chongfu = redis_rc.get('danao_fugai_chongfu')
            chongfu = 0
            if fugai_chongfu:
                chongfu = fugai_chongfu
            # 分页
            if length != 0:
                start_line = (current_page - 1) * length
                stop_line = start_line + length
                objs = objs[start_line: stop_line]
            # 返回的数据
            data_list = []
            for obj in objs:
                if obj.json_detail_data:
                    pass
                if str(obj.search_engine) == '1':
                    yinqing = '百度'
                elif str(obj.search_engine) == '4':
                    yinqing = '手机百度'
                elif str(obj.search_engine) == '3':
                    yinqing = '360'
                elif str(obj.search_engine) == '6':
                    yinqing = '手机360'
                else:
                    yinqing = ''
                rank_num = 0
                paiming_detail = '-'
                if obj.paiming_detail:
                    if len(obj.paiming_detail) == 1:
                        paiming_detail = int(obj.paiming_detail)
                        if int(paiming_detail) > 0:
                            rank_num =1
                    if len(obj.paiming_detail) > 1:
                        rank_num = len(obj.paiming_detail.split(','))
                        paiming_detail_sort =sorted(set(eval(obj.paiming_detail)))
                        if type(paiming_detail_sort) != int:
                            ls2 = [str(i) for i in paiming_detail_sort]
                            paiming_detail = ','.join(ls2)
                data_list.append({
                    'id':obj.id,
                    'keyword':obj.keyword,
                    'search_engine':yinqing,
                    'rank_info':paiming_detail,
                    'otherData':obj.json_detail_data,
                    'rank_num':rank_num
                    })

            response.code = 200
            response.msg = '查询成功'
            response.data = {
                'retData': data_list,                   # 详情
                'dataCount': dataCount,                 # 任务总数
                'paiming_num':fugai_num,           # 有排名数
                'fugailv':fugailv,                      # 覆盖率
                'paiminglv':paiminglv,                  # 排名率
                'yiwancheng_obj':yiWanCheng,            # 已完成数量
                'chongfu_num':int(chongfu),             # 重复数
                'whether_complete':whether_complete,    # 是否全部完成
                'query_progress':query_progress         # 进度条
            }
        else:
            response.code = 402
            response.msg = "请求异常"
            response.data = json.loads(forms_obj.errors.as_json())

    return JsonResponse(response.__dict__)


#  增删改
#  csrf  token验证
@csrf_exempt
@account.is_token(models.zhugedanao_userprofile)
def fuGaiChaXun(request, oper_type, o_id):
    response = Response.ResponseObj()
    user_id = request.GET.get('user_id')
    if request.method == "POST":
        # 增加收录任务
        if oper_type == "add":
            form_data = {
                'search_list': request.POST.get('searchEngineModel'),
                'keywords_list': request.POST.get('editor_content'),
                'conditions_list':request.POST.get('fugai_tiaojian')
            }
            models.zhugedanao_fugai_chaxun.objects.filter(user_id_id=user_id).delete()
            #  创建 form验证 实例（参数默认转成字典）
            forms_obj = AddForm(form_data)
            if forms_obj.is_valid():
                chongfu = int(len(forms_obj.cleaned_data.get('keywords_list'))) - int(len(set(forms_obj.cleaned_data.get('keywords_list'))))
                redis_rc.set('danao_fugai_chongfu', '{}'.format(int(chongfu)), ex=None, px=None, nx=False, xx=False)

                print(redis_rc.get('danao_fugai_chongfu'))
                 # 添加数据库
                search_list = forms_obj.cleaned_data.get('search_list')
                keywords_list = set(forms_obj.cleaned_data.get('keywords_list'))
                conditions_list = forms_obj.cleaned_data.get('conditions_list')
                querysetlist = []
                now_date = datetime.datetime.today().strftime('%Y-%m-%d %H:%M:%S')
                for search in json.loads(search_list):
                    for keyword in list(keywords_list):
                        querysetlist.append(
                            models.zhugedanao_fugai_chaxun(
                                user_id_id=user_id,
                                keyword=keyword,
                                search_engine=search,
                                sousuo_guize=','.join(conditions_list),
                                createAndStart_time=now_date
                            )
                        )
                models.zhugedanao_fugai_chaxun.objects.bulk_create(querysetlist)
                response.code = 200
                response.msg = "删除原数据与添加新数据成功"
            else:
                print("验证不通过")
                response.code = 301
                response.msg = json.loads(forms_obj.errors.as_json())


    elif request.method == 'GET':
        # 点击返回 删除任务
        if oper_type == 'clickReturn':
            models.zhugedanao_fugai_chaxun.objects.filter(user_id_id=user_id).delete()
            response.code = 200
            response.msg = "退出成功"
            return JsonResponse(response.__dict__)

        # 生成报表
        if oper_type == 'generateExcel':
            wb = Workbook()
            ws = wb.active
            ws.title = '关键词覆盖查询'
            ws.cell(row=1, column=1, value="关键词覆盖查询")
            ws.cell(row=2, column=3, value="查询时间:")
            ws.cell(row=8, column=1, value="关键词")
            ws.cell(row=8, column=2, value="排名个数")
            ws.cell(row=8, column=3, value="排名情况")
            ws.cell(row=8, column=4, value="搜索引擎")
            ft1 = Font(name='宋体', size=22)
            a1 = ws['A1']
            a1.font = ft1

            # # 合并单元格        开始行      结束行       用哪列          占用哪列
            ws.merge_cells(start_row=1, end_row=1, start_column=1, end_column=4)
            ws.merge_cells(start_row=2, end_row=7, start_column=1, end_column=1)
            ws.merge_cells(start_row=2, end_row=7, start_column=2, end_column=2)
            ws.merge_cells(start_row=2, end_row=7, start_column=3, end_column=3)
            ws.merge_cells(start_row=2, end_row=7, start_column=4, end_column=4)
            # ws.merge_cells(start_row=2, end_row=5, start_column=5, end_column=5)

            # print('设置列宽')
            ws.column_dimensions['A'].width = 35
            ws.column_dimensions['B'].width = 13
            ws.column_dimensions['C'].width = 30
            ws.column_dimensions['D'].width = 30

            # # print('设置行高')
            ws.row_dimensions[1].height = 28

            # # print('文本居中')
            ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
            ws['C2'].alignment = Alignment(horizontal='right', vertical='center')
            ws['A8'].alignment = Alignment(horizontal='center', vertical='center')
            ws['B8'].alignment = Alignment(horizontal='center', vertical='center')
            ws['C8'].alignment = Alignment(horizontal='center', vertical='center')
            ws['D8'].alignment = Alignment(horizontal='center', vertical='center')
            ws['D2'].alignment = Alignment(horizontal='left', vertical='center')

            ws2 = wb.create_sheet('sheet2')
            ws2.title = '关键词覆盖查询详情'
            ws2.cell(row=1, column=1, value="关键词覆盖查询详情")
            ws2.cell(row=2, column=4, value="查询时间:")
            ws2.cell(row=3, column=1, value="关键词")
            ws2.cell(row=3, column=2, value="名次")
            ws2.cell(row=3, column=3, value="标题")
            ws2.cell(row=3, column=4, value="链接")
            ws2.cell(row=3, column=5, value="规则")
            ws2.cell(row=3, column=6, value="搜索引擎")
            ft1 = Font(name='宋体', size=22)
            a1 = ws2['A1']
            a1.font = ft1

            # # 合并单元格        开始行      结束行       开始列          结束列
            ws2.merge_cells(start_row=1, end_row=1, start_column=1, end_column=6)
            ws2.merge_cells(start_row=2, end_row=2, start_column=4, end_column=5)

            # # print('设置列宽')
            ws2.column_dimensions['A'].width = 35
            ws2.column_dimensions['B'].width = 15
            ws2.column_dimensions['C'].width = 60
            ws2.column_dimensions['D'].width = 30
            ws2.column_dimensions['E'].width = 13
            ws2.column_dimensions['F'].width = 20
            #
            # # print('设置行高')
            ws2.row_dimensions[1].height = 80
            ws2.row_dimensions[2].height = 30
            #
            # # print('文本居中')
            ws2['A1'].alignment = Alignment(horizontal='center', vertical='center')
            ws2['A3'].alignment = Alignment(horizontal='center', vertical='center')
            ws2['B3'].alignment = Alignment(horizontal='center', vertical='center')
            ws2['C3'].alignment = Alignment(horizontal='center', vertical='center')
            ws2['D3'].alignment = Alignment(horizontal='center', vertical='center')
            ws2['E3'].alignment = Alignment(horizontal='center', vertical='center')
            ws2['F3'].alignment = Alignment(horizontal='center', vertical='center')
            ws2['C2'].alignment = Alignment(horizontal='center', vertical='center')
            ws2['D2'].alignment = Alignment(horizontal='right', vertical='center')
            ws2['F2'].alignment = Alignment(horizontal='left', vertical='center')

            objs = models.zhugedanao_fugai_chaxun.objects.filter(user_id_id=user_id)
            row = 9
            row_two = 4
            for obj in objs:
                if str(obj.search_engine) == '1':
                    yinqing = '百度'
                elif str(obj.search_engine) == '4':
                    yinqing = '手机百度'
                elif str(obj.search_engine) == '3':
                    yinqing = '360'
                elif str(obj.search_engine) == '6':
                    yinqing = '手机360'
                else:
                    yinqing = ''
                paiming_num = 0
                if obj.paiming_detail:
                    paiming_detail = obj.paiming_detail.split(',')
                    paiming_num = len(paiming_detail)
                ws.cell(row=2, column=4, value="{chaxun_time}".format(chaxun_time=obj.createAndStart_time))
                ws2.cell(row=2, column=6, value="{chaxun_time}".format(chaxun_time=obj.createAndStart_time))
                ws.cell(row=row, column=1, value="{keyword}".format(keyword=obj.keyword))
                ws.cell(row=row, column=2, value="{paming_num}".format(paming_num=paiming_num))
                ws.cell(row=row, column=3, value="{paiming_detail}".format(paiming_detail=obj.paiming_detail))
                ws.cell(row=row, column=4, value="{search}".format(search=yinqing))
                row += 1
                if obj.json_detail_data:
                    for data_detail in json.loads(obj.json_detail_data):
                        ws2.cell(row=row_two, column=1, value="{keyword}".format(keyword=data_detail['keyword']))
                        ws2.cell(row=row_two, column=2, value="{paiming}".format(paiming=data_detail['rank']))
                        ws2.cell(row=row_two, column=3, value="{title}".format(title=data_detail['title']))
                        ws2.cell(row=row_two, column=4, value="{title_url}".format(title_url=data_detail['url']))
                        ws2.cell(row=row_two, column=5, value="{guize}".format(guize=data_detail['guize']))
                        ws2.cell(row=row_two, column=6, value="{search}".format(search=data_detail['search_engine']))
                        row_two += 1

            randInt = random.randint(1, 100)
            nowDateTime = int(time.time())
            excel_name = str(randInt) + str(nowDateTime)
            download_excel_path = 'http://api.zhugeyingxiao.com/' + os.path.join('statics', 'zhugedanao', 'fuGaiExcel', '{}.xlsx'.format(excel_name))
            wb.save(os.path.join(os.getcwd(), 'statics', 'zhugedanao', 'fuGaiExcel', '{}.xlsx'.format(excel_name)))
            response.code = 200
            response.msg = '生成成功'
            # download_excel_path = 'http://127.0.0.1:8000/' + os.path.join('statics', 'zhugedanao', 'fuGaiExcel', '{}.xlsx'.format(excel_name))
            response.data = {'excel_name': download_excel_path}
            return JsonResponse(response.__dict__)

    else:
        response.code = 402
        response.msg = "请求异常"


    return JsonResponse(response.__dict__)