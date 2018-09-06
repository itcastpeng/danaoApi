
from openpyxl.styles import Font, Alignment
from openpyxl import Workbook
from zhugedanao import models
from publicFunc import Response
from publicFunc import account
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt, csrf_protect
import datetime
from zhugedanao.forms.guanjianci_paiming_chaxun import AddForm, SelectForm
import json, time, re, os, sys
import random
# cerf  token验证 用户展示模块
@csrf_exempt
@account.is_token(models.zhugedanao_userprofile)
def guanJianCiPaiMingShow(request):
    response = Response.ResponseObj()
    user_id = request.GET.get('user_id')
    if request.method == "GET":
        forms_obj = SelectForm(request.GET)
        if forms_obj.is_valid():
            current_page = forms_obj.cleaned_data['current_page']
            length = forms_obj.cleaned_data['length']
            task_list_objs = models.zhugedanao_guanjianci_paiming_chaxun.objects.filter(
                user_id_id=user_id)
            yiwancheng_count = task_list_objs.filter(user_id_id=user_id).filter(is_perform=1).count()
            print('yiwancheng_count======> ',yiwancheng_count)
            obj_count = task_list_objs.count()
            yiwancheng_obj = 0
            query_progress = 0
            whether_complete = False
            if yiwancheng_count:
                yiwancheng_obj = int(yiwancheng_count)
                query_progress = int((int(yiwancheng_count) / int(obj_count)) * 100)
                if int(yiwancheng_obj) == int(obj_count):
                    whether_complete = True
            print('yiwancheng_obj=====> ',yiwancheng_obj, query_progress)
            chongfu_num = 0
            # 分页
            if length != 0:
                start_line = (current_page - 1) * length
                stop_line = start_line + length
                objs = task_list_objs[start_line: stop_line]
            data_list = []
            for obj in objs:
                yinqing = '百度'
                if int(obj.search_engine) == 1:
                    yinqing = '百度'
                elif int(obj.search_engine) == 4:
                    yinqing = '手机百度'
                elif int(obj.search_engine) == 3:
                    yinqing = '360'
                elif int(obj.search_engine) == 6:
                    yinqing = '手机360'
                paiming = '-'
                if obj.paiming:
                    paiming = int(obj.paiming)
                data_list.append({
                    'paiming':paiming,
                    'lianjie':obj.lianjie,
                    'keyword':obj.keyword,
                    'search':yinqing,
                })
            other_data = {
                'query_progress': query_progress,
                'whether_complete': whether_complete,
                'yiwancheng_obj': yiwancheng_obj,
                'data_list':data_list,
                'chongfu_num': chongfu_num,
                'obj_count': obj_count,
            }
            response.msg = '查询成功'
            response.code = 200
            response.data = {'other_data':other_data}
        else:
            response.code = 402
            response.msg = "数据类型验证失败"
            response.data = json.loads(forms_obj.errors.as_json())
    return JsonResponse(response.__dict__)


#  增删改
#  csrf  token验证
@csrf_exempt
@account.is_token(models.zhugedanao_userprofile)
def guanJianCiPaiMingOper(request, oper_type, o_id):
    response = Response.ResponseObj()
    user_id = request.GET.get('user_id')
    if request.method == "POST":
        # 增加重点词任务
        if oper_type == "add":
            user_id = request.GET.get('user_id')
            form_data = {
                'search_engine' : request.POST.get('search_engine'),
                'keywords' : request.POST.get('keywords'),
            }
            forms_obj = AddForm(form_data)
            if forms_obj.is_valid():
                models.zhugedanao_guanjianci_paiming_chaxun.objects.filter(user_id_id=user_id).delete()
                keywords = forms_obj.cleaned_data.get('keywords')
                search_engine = forms_obj.cleaned_data.get('search_engine')
                # keyword_list = list(set(keywords.split('\n')))        # 去重
                keyword_list = list(keywords.split('\n'))
                querysetlist = []
                for search in search_engine:
                    num = 0
                    for keywords in keyword_list:
                        num += 1
                        if keywords:
                            if 'http' in keywords:
                                re_keyword = re.findall("(.*)http", keywords.replace('\t', ''))
                                if re_keyword[0]:
                                    keyword = re_keyword[0]
                                    url_list = keywords.split(keyword)
                                    url = url_list[1]
                                    if url and keyword:
                                        pattern = re.compile(
                                            r'http[s]?://(?:[a-zA-Z]|[0-9]|[$-_@.&+]|[!*\(\),]|(?:%[0-9a-fA-F][0-9a-fA-F]))+')  # 判断链接
                                        url = re.findall(pattern, url)
                                        if url[0]:
                                            print(url[0], keyword, search)
                                            response.code = 200
                                            response.msg = '添加成功'
                                            querysetlist.append(
                                                models.zhugedanao_guanjianci_paiming_chaxun(
                                                    search_engine=search,
                                                    keyword=keyword,
                                                    lianjie=url[0],
                                                    user_id_id=user_id
                                                )
                                            )
                                        else:
                                            response.code = 301
                                            response.msg = '第{}行请输入正确链接!'.format(num)
                                            break
                                    else:
                                        response.code = 301
                                        response.msg = '第{}行请输入正确链接!'.format(num)
                                        break
                                else:
                                    response.code = 301
                                    response.msg = '第{}行请输入关键词!'.format(num)
                                    break
                            else:
                                response.code = 301
                                response.msg = '第{}行请输入正确链接!'.format(num)
                                break
                models.zhugedanao_guanjianci_paiming_chaxun.objects.bulk_create(querysetlist)
            else:
                print("验证不通过")
                response.code = 301
                response.msg = json.loads(forms_obj.errors.as_json())
            response.data = {}


    elif request.method == 'GET':

        # 退出
        if oper_type == 'clickReturn':
            models.zhugedanao_guanjianci_paiming_chaxun.objects.filter(user_id_id=user_id).delete()
            response.code = 200
            response.msg = '退出成功'
            return JsonResponse(response.__dict__)

        # 生成excel表格
        if oper_type == 'generateExcel':
            now_date = datetime.date.today().strftime('%Y-%m-%d')
            wb = Workbook()
            ws = wb.active
            ws.cell(row=1, column=1, value="序号").font = Font(b=True)
            ws.cell(row=1, column=2, value="关键词").font = Font(b=True)
            ws.cell(row=1, column=3, value="网址").font = Font(b=True)
            ws.cell(row=1, column=4, value="搜索引擎").font = Font(b=True)
            ws.cell(row=1, column=5, value="排名").font = Font(b=True)
            ws.cell(row=1, column=6, value="制表日期:{}".format(now_date)).font = Font(b=True)

            # 合并单元格        开始行      结束行       用哪列          占用哪列
            # ws.merge_cells(start_row=1, end_row=2, start_column=1, end_column=2)

            # print('设置列宽')
            ws.column_dimensions['A'].width = 8
            ws.column_dimensions['B'].width = 20
            ws.column_dimensions['C'].width = 20
            ws.column_dimensions['D'].width = 10
            ws.column_dimensions['E'].width = 8
            ws.column_dimensions['F'].width = 15

            # print('设置行高')
            ws.row_dimensions[1].height = 30
            ws.row_dimensions[2].height = 30
            ws.row_dimensions[3].height = 30
            ws.row_dimensions[4].height = 30
            ws.row_dimensions[5].height = 30
            ws.row_dimensions[6].height = 30

            # print('文本居中')
            ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
            ws['B1'].alignment = Alignment(horizontal='center', vertical='center')
            ws['C1'].alignment = Alignment(horizontal='center', vertical='center')
            ws['D1'].alignment = Alignment(horizontal='center', vertical='center')
            ws['E1'].alignment = Alignment(horizontal='center', vertical='center')
            ws['F1'].alignment = Alignment(horizontal='center', vertical='center')
            row = 2
            objs = models.zhugedanao_guanjianci_paiming_chaxun.objects.filter(user_id_id=user_id)
            yinqing = '百度'
            number = 0
            for obj in objs:
                ws.row_dimensions[int(row)+5].height = 30
                if int(obj.search_engine) == 1:
                    yinqing = '百度'
                elif int(obj.search_engine) == 4:
                    yinqing = '手机百度'
                elif int(obj.search_engine) == 3:
                    yinqing = '360'
                elif int(obj.search_engine) == 6:
                    yinqing = '手机360'
                paiming = '-'
                if obj.paiming:
                    paiming = int(obj.paiming)
                number += 1
                number_ws = ws.cell(row=row, column=1, value="{number}".format(number=number))
                number_ws.alignment = Alignment(horizontal='center', vertical='center')
                number_ws.font = Font(size=10)
                keyword_ws = ws.cell(row=row, column=2, value="{title}".format(title=obj.keyword))
                keyword_ws.font = Font(size=10)
                keyword_ws.alignment = Alignment(vertical='center')

                lianjie_ws = ws.cell(row=row, column=3, value="{lianjie}".format(lianjie=obj.lianjie))
                lianjie_ws.font = Font(size=10)

                search_ws = ws.cell(row=row, column=4, value="{search}".format(search=yinqing))
                search_ws.alignment = Alignment(horizontal='center', vertical='center')
                search_ws.font = Font(size=10)

                search_ws = ws.cell(row=row, column=5, value="{paiming}".format(paiming=paiming))
                search_ws.alignment = Alignment(vertical='center')
                search_ws.font = Font(size=10)
                search_ws.alignment = Alignment(horizontal='center', vertical='center')
                row += 1

            paiming_num = objs.filter(paiming__isnull=False).count()  # 排名总数
            num_count_ws = ws.cell(row=2, column=6, value="数据总数：{}".format(objs.count()))
            num_count_ws.font = Font('宋体', color='0066CD', size=10, b=True)
            num_count_ws.alignment = Alignment(vertical='center')
            chongfu_ws = ws.cell(row=3, column=6, value="重复：{}".format(0))
            chongfu_ws.font = Font('宋体', color='0066CD', size=10, b=True)
            chongfu_ws.alignment = Alignment(vertical='center')

            baidu_pc_paiming_num = objs.filter(search_engine=1).filter(paiming__isnull=False).count() # 百度pc排名总数
            pc_bili = 0
            if baidu_pc_paiming_num:
                pc_bili = int((int(baidu_pc_paiming_num) / int(paiming_num)) * 100)
            pcpaiming_ws = ws.cell(row=4, column=6, value="百度排名数：{}".format(baidu_pc_paiming_num))
            pcpaiming_ws.font = Font('宋体', color='0066CD', size=10, b=True)
            pcpaiming_ws.alignment = Alignment(vertical='center')
            pcpaimingbili_ws = ws.cell(row=5, column=6, value="百度排名比例：{}".format(pc_bili))
            pcpaimingbili_ws.font = Font('宋体', color='0066CD', size=10, b=True)
            pcpaimingbili_ws.alignment = Alignment(vertical='center')

            baidu_mobiel_paiming_num = objs.filter(search_engine=1).filter(paiming__isnull=False).count()   # 百度移动排名总数
            mobiel_bili = 0
            if baidu_mobiel_paiming_num:
                mobiel_bili = int((int(baidu_mobiel_paiming_num) / int(paiming_num)) * 100)
            mbpaiming_ws = ws.cell(row=6, column=6, value="手机百度排名数：{}".format(baidu_mobiel_paiming_num))
            mbpaiming_ws.font = Font('宋体', color='0066CD', size=10, b=True)
            mbpaiming_ws.alignment = Alignment(vertical='center')
            mbpaimingbili_ws = ws.cell(row=7, column=6, value="手机百度排名比例：{}".format(mobiel_bili))
            mbpaimingbili_ws.font = Font('宋体', color='0066CD', size=10, b=True)
            mbpaimingbili_ws.alignment = Alignment(vertical='center')

            randInt = random.randint(1, 100)
            nowDateTime = int(time.time())
            excel_name = str(randInt) + str(nowDateTime)
            wb.save(os.path.join(os.getcwd(), 'statics', 'zhugedanao', 'guanjianci_paiming', '{}.xlsx'.format(excel_name)))
            response.code = 200
            response.msg = '生成成功'
            response.data = {'excel_name': 'http://api.zhugeyingxiao.com/' + os.path.join('statics', 'zhugedanao', 'guanjianci_paiming','{}.xlsx'.format(excel_name))}
            return JsonResponse(response.__dict__)

    else:
        response.code = 402
        response.msg = "请求异常"
    return JsonResponse(response.__dict__)