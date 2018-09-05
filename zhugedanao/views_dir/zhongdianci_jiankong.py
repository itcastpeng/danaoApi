
from openpyxl.styles import Font, Alignment
from openpyxl import Workbook
from zhugedanao import models
from publicFunc import Response
from publicFunc import account
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt, csrf_protect
import datetime
from zhugedanao.forms.zhongdianci_jiankong import AddForm, SelectForm
import json, time, re, os, sys

# cerf  token验证 用户展示模块
@csrf_exempt
@account.is_token(models.zhugedanao_userprofile)
def zhongDianCiShowTaskList(request):
    response = Response.ResponseObj()
    user_id = request.GET.get('user_id')
    if request.method == "GET":
        forms_obj = SelectForm(request.GET)
        if forms_obj.is_valid():
            current_page = forms_obj.cleaned_data['current_page']
            length = forms_obj.cleaned_data['length']
            task_list_objs = models.zhugedanao_zhongdianci_jiankong_taskList.objects.filter(
                user_id_id=user_id)
            # 分页
            if length != 0:
                start_line = (current_page - 1) * length
                stop_line = start_line + length
                objs = task_list_objs[start_line: stop_line]
            if objs:
                data_list = []
                for obj in objs:
                    # 查询跑出来的任务 数量 判断百分比
                    # detail_objs_count = obj.zhugedanao_zhongdianci_jiankong_taskdetail_set.filter(
                    #     tid_id=obj.id,
                    # )
                    # detail_count = detail_objs_count.filter(is_perform=0).filter(task_start_time__isnull=False).count()
                    # baifenbi = 0
                    # if detail_count:
                    #     baifenbi = int((detail_count / detail_objs_count.count()) * 100)
                    # obj.task_jindu = baifenbi
                    # obj.save()
                    qiyongstatus = '未启用'
                    if obj.qiyong_status:
                        qiyongstatus = '已启用'
                    task_status = '正在查询'

                    if int(obj.task_status) == 2:
                        task_status = '未查询'
                    elif int(obj.task_status) == 1:
                        task_status = '已查询'
                    data_list.append({
                        "id": obj.id,
                        'zhixing':obj.is_zhixing,
                        "qiyong_status": qiyongstatus,
                        "task_name": obj.task_name,
                        "task_start_time": obj.task_start_time,
                        "task_status": task_status,
                        "search_engine": obj.search_engine.split(','),
                        'task_jindu': obj.task_jindu,
                    })
                    if int(obj.task_jindu) == 100:
                        task_list_objs.filter(id=obj.id).update(
                            task_status=1,
                            is_zhixing=0
                        )
                        next_datetime = obj.next_datetime
                        now_date_start = obj.task_start_time
                        now_date = datetime.date.today().strftime('%Y-%m-%d')  # 当前年月日
                        canshu = now_date + ' ' + now_date_start
                        next_datetime_start = datetime.datetime.today().strptime(canshu, "%Y-%m-%d %H:%M:%S")  # 传来的参数 时分秒
                        now_date = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                        now_datetime = datetime.datetime.strptime(now_date, '%Y-%m-%d %H:%M:%S')
                        print(next_datetime, now_datetime)
                        if next_datetime <= now_datetime:
                            next_datetime_addoneday = (next_datetime_start + datetime.timedelta(days=1)).strftime(
                                '%Y-%m-%d %H:%M:%S')
                            models.zhugedanao_zhongdianci_jiankong_taskList.objects.filter(id=obj.id).update(
                                next_datetime=next_datetime_addoneday,
                                is_zhixing=0
                            )
                response.data = {'data_list':data_list}
            else:
                data_list = []
                response.data = {'data_list':data_list}
            response.msg = '查询成功'
            response.code = 200
        else:
            response.code = 402
            response.msg = "数据类型验证失败"
            response.data = json.loads(forms_obj.errors.as_json())
    return JsonResponse(response.__dict__)


@csrf_exempt
@account.is_token(models.zhugedanao_userprofile)
def zhongDianCiDetailShowTaskList(request):
    response = Response.ResponseObj()
    user_id = request.GET.get('user_id')
    tid = request.GET.get('tid')
    forms_obj = SelectForm(request.GET)
    exit_data_list = {}
    if forms_obj.is_valid():
        current_page = forms_obj.cleaned_data['current_page']
        length = forms_obj.cleaned_data['length']
        # print('tid=============> ',tid)
        objs = models.zhugedanao_zhongdianci_jiankong_taskDetail.objects.select_related('tid__user_id').filter(
            tid__user_id=user_id,
            tid=tid
        )
        detail_task_count = objs.count()
        # 分页
        if length != 0:
            start_line = (current_page - 1) * length
            stop_line = start_line + length
            objs = objs[start_line: stop_line]
        data_list = []
        headers_list = []
        if objs:
            for obj in objs:
                detail_objs = models.zhugedanao_zhongdianci_jiankong_taskDetailData.objects.filter(tid_id=obj.id)[0:3]
                sanci_chaxun = {}
                if detail_objs:
                    for detail_obj in detail_objs:
                        shoulu = '否'
                        if detail_obj.is_shoulu == 1:
                            shoulu = '是'
                        detail_create = detail_obj.create_time.strftime('%Y-%m-%d')
                        if detail_create not in headers_list:
                            headers_list.append(str(detail_create))
                        sanci_chaxun[detail_create] = {
                            'detail_create': detail_create,
                            'shoulu': shoulu,
                            'paiming': detail_obj.paiming
                        }
                data_list.append({
                    "id": obj.id,
                    # "tid": obj.tid_id,
                    "search_engine": obj.search_engine,
                    "lianjie": obj.lianjie,
                    "keywords": obj.keyword,
                    "mohupipei": obj.mohupipei,
                    # "create_time": obj.create_time,
                    'sanci_chaxun':sanci_chaxun,
                })
                # print('data_list====> ',data_list)
            exit_data_list = {
                'count_page':detail_task_count,
                'data_list':data_list,
                'task_name':objs[0].tid.task_name,
                'headers_list':headers_list,
            }
            response.data = json.dumps(exit_data_list)
        else:
            response.data = {}
        response.code = 200
        response.msg = '查询成功'
    else:
        response.code = 402
        response.msg = "数据类型验证失败"
        response.data = json.loads(forms_obj.errors.as_json())
    return JsonResponse(response.__dict__)

#  增删改
#  csrf  token验证
@csrf_exempt
@account.is_token(models.zhugedanao_userprofile)
def zhongDianCiOper(request, oper_type, o_id):
    response = Response.ResponseObj()
    user_id = request.GET.get('user_id')
    if request.method == "POST":
        # 增加重点词任务
        if oper_type == "add":
            user_id = request.GET.get('user_id')
            qiyong_status =  request.POST.get('qiyong_status'),
            form_data = {
                'mohupipei' : request.POST.get('mohupipei'),
                'task_name' : request.POST.get('task_name'),
                'search_engine' : request.POST.get('search_engine'),
                'keywords' : request.POST.get('keywords'),
                'task_start_time' : request.POST.get('task_start_time'),
                'task_status' : request.POST.get('task_status'),
            }
            forms_obj = AddForm(form_data)
            if forms_obj.is_valid():
                keywords = forms_obj.cleaned_data.get('keywords')
                # keyword_list = list(set(keywords.split('\n')))        # 去重
                keyword_list = list(keywords.split('\n'))
                qiyongstatus = False
                if qiyong_status:
                    qiyongstatus = True
                now_date = datetime.date.today().strftime('%Y-%m-%d')  # 当前年月日
                canshu = now_date + ' ' + form_data['task_start_time']
                kaishishijian = datetime.datetime.today().strptime(canshu, "%Y-%m-%d %H:%M:%S")  # 传来的参数 时分秒
                now = now_date + ' ' + time.strftime("%H:%M:%S")
                now_time = datetime.datetime.today().strptime(now, "%Y-%m-%d %H:%M:%S")  # 当前时分秒
                next_datetime = kaishishijian
                if kaishishijian < now_time:
                    next_datetime = (kaishishijian + datetime.timedelta(days=1)).strftime('%Y-%m-%d %H:%M:%S')
                objs = models.zhugedanao_zhongdianci_jiankong_taskList.objects.create(
                    task_name=forms_obj.cleaned_data.get('task_name'),
                    search_engine=','.join(forms_obj.cleaned_data.get('search_engine')),
                    next_datetime=next_datetime,
                    task_start_time=form_data['task_start_time'],
                    qiyong_status=qiyongstatus,
                    user_id_id=user_id
                )
                querysetlist = []
                create_time = datetime.datetime.today().strftime("%Y-%m-%d %H-%M-%S")
                flag = False
                panduan_flag = False
                for search in forms_obj.cleaned_data.get('search_engine'):
                    if panduan_flag:
                        break
                    num = 0
                    panduan_number = 1
                    for keywords in keyword_list:
                        num += 1
                        if keywords:
                            print(keywords)
                            if 'http' in keywords:
                                re_keyword = re.findall("(.*)http", keywords.replace('\t', ''))
                                if re_keyword[0]:
                                    url_list = keywords.split(re_keyword[0])
                                    url = url_list[1]
                                else:
                                    response.code = 301
                                    response.msg = '第{}行请输入关键词!'.format(num)
                                    panduan_flag = True
                                    flag = True
                                    break
                                if url and re_keyword:
                                    pattern = re.compile(
                                        r'http[s]?://(?:[a-zA-Z]|[0-9]|[$-_@.&+]|[!*\(\),]|(?:%[0-9a-fA-F][0-9a-fA-F]))+')  # 判断链接
                                    url = re.findall(pattern, url)
                                    if url:
                                        if panduan_number >= 1:
                                            if panduan_number != num:
                                                panduan_flag = True
                                                response.code = 301
                                                flag = True
                                                response.msg = '第{}行请输入正确链接'.format(num - 1)
                                                break
                                        panduan_number += 1
                                        querysetlist.append(
                                            models.zhugedanao_zhongdianci_jiankong_taskDetail(
                                                tid_id=objs.id,
                                                search_engine=search,
                                                lianjie=url[0],
                                                keyword=re_keyword[0],
                                                create_time=create_time
                                            )
                                        )
                                        response.code = 200
                                        response.msg = "添加成功"
                                    else:
                                        flag = True
                                        response.code = 301
                                        response.msg = '第{}行请输入正确链接!'.format(num)
                                else:
                                    flag = True
                                    response.code = 301
                                    if not url:
                                        response.msg = '第{}行请填写正确链接!'.format(num)
                                    if not re_keyword:
                                        response.msg = '第{}行请填写正确关键词!'.format(num)
                            else:
                                if form_data['mohupipei']:
                                    models.zhugedanao_zhongdianci_jiankong_taskList.objects.filter(id=objs.id).update(
                                        mohupipei=form_data['mohupipei']
                                    )
                                    querysetlist.append(
                                        models.zhugedanao_zhongdianci_jiankong_taskDetail(
                                            tid_id=objs.id,
                                            search_engine=search,
                                            keyword=keywords.strip(),
                                            mohupipei=form_data['mohupipei'],
                                            create_time=create_time
                                        )
                                    )
                                    response.code = 200
                                    response.msg = "添加成功"
                                else:
                                    flag = True
                                    response.code = 301
                                    response.msg = '无链接, 请填写模糊匹配或填写链接!'
                        else:
                            flag = True
                            response.code = 301
                            response.msg = '第{}行不能为空!'.format(num)
                models.zhugedanao_zhongdianci_jiankong_taskDetail.objects.bulk_create(querysetlist)
                if flag:
                    # 验证不通过 删除创建的任务列表
                    task_objs = models.zhugedanao_zhongdianci_jiankong_taskList.objects
                    task_objs.filter(id=objs.id)
                    if task_objs:
                        task_objs.filter(id=objs.id).delete()
            else:
                print("验证不通过")
                response.code = 301
                response.msg = json.loads(forms_obj.errors.as_json())

        # 清空任务
        if oper_type == 'empty':
            detail_objs = models.zhugedanao_zhongdianci_jiankong_taskDetail.objects.filter(tid=o_id)
            for detail_obj in detail_objs:
                detail_obj.zhugedanao_zhongdianci_jiankong_taskdetaildata_set.filter(tid=detail_obj.id).delete()
            # detail_objs.delete()
            response.msg = '清空成功'
            response.code = 200
            return JsonResponse(response.__dict__)

        # 删除任务
        if oper_type == 'exct_delete':
            id_list = request.POST.get('id_list')
            json_id_list = json.loads(id_list)
            for id in json_id_list:
                detail_objs = models.zhugedanao_zhongdianci_jiankong_taskDetail.objects.filter(tid=id)
                for detail_obj in detail_objs:
                    detail_obj.zhugedanao_zhongdianci_jiankong_taskdetaildata_set.filter(tid=detail_obj.id).delete()
                objs = models.zhugedanao_zhongdianci_jiankong_taskList.objects.filter(
                    user_id_id=user_id,
                    id=id
                )
                objs.delete()
                detail_objs.delete()
                response.msg = '删除成功'
                response.code = 200


        # 查询修改前信息
        if oper_type == 'update_show':
            objs = models.zhugedanao_zhongdianci_jiankong_taskList.objects.filter(
                user_id_id=user_id,
                id=o_id
            )
            if objs:
                task_name = objs[0].task_name
                task_start_time = objs[0].task_start_time
                qiyong_status = objs[0].qiyong_status
                response.code = 200
                response.msg = '修改查询成功'
                response.data = {
                    'task_name':task_name,
                    'task_start_time':task_start_time,
                    'qiyong_status':qiyong_status
                }
                return JsonResponse(response.__dict__)

        # 确认修改
        if oper_type == 'update':
            task_name = request.POST.get('task_name')
            task_start_time = request.POST.get('task_start_time')
            qiyong_status = request.POST.get('qiyong_status')
            print('task_name=-======> ',task_name , task_start_time)
            objs = models.zhugedanao_zhongdianci_jiankong_taskList.objects.filter(
                user_id_id=user_id,
                id=o_id
            )

            now_date = datetime.date.today().strftime('%Y-%m-%d')  # 当前年月日
            canshu = now_date + ' ' + task_start_time
            kaishishijian = datetime.datetime.today().strptime(canshu, "%Y-%m-%d %H:%M:%S")  # 传来的参数 时分秒
            now = now_date + ' ' + time.strftime("%H:%M:%S")
            now_time = datetime.datetime.today().strptime(now, "%Y-%m-%d %H:%M:%S")  # 当前时分秒
            next_datetime = kaishishijian
            if kaishishijian < now_time:
                next_datetime = (kaishishijian + datetime.timedelta(days=1)).strftime('%Y-%m-%d %H:%M:%S')
            objs.update(
                task_name=task_name,
                task_start_time=task_start_time,
                next_datetime=next_datetime,
                qiyong_status=qiyong_status
            )
            response.code = 200
            response.msg = '修改成功'

            return JsonResponse(response.__dict__)


    elif request.method == 'GET':
        # 生成excel表格
        if oper_type == 'generateExcel':
            now_date = datetime.datetime.today().strftime('%Y-%m-%d %H-%M-%S')
            wb = Workbook()
            ws = wb.active
            ws.cell(row=1, column=1, value="重点关键词")
            ws.cell(row=2, column=1, value="任务名称:")
            ws.cell(row=3, column=1, value="关键词")
            ws.cell(row=3, column=2, value="网址")
            ws.cell(row=3, column=3, value="搜索引擎")
            ws.cell(row=4, column=4, value="排名")
            ws.cell(row=4, column=5, value="排名")
            ws.cell(row=4, column=6, value="排名")
            ft1 = Font(name='宋体', size=22)
            a1 = ws['A1']
            a1.font = ft1

            # 合并单元格        开始行      结束行       哪列做改变       占几列
            ws.merge_cells(start_row=1, end_row=1, start_column=1, end_column=6)
            ws.merge_cells(start_row=3, end_row=4, start_column=1, end_column=1)
            ws.merge_cells(start_row=3, end_row=4, start_column=2, end_column=2)
            ws.merge_cells(start_row=3, end_row=4, start_column=3, end_column=3)

            # print('设置列宽')
            ws.column_dimensions['A'].width = 50
            ws.column_dimensions['B'].width = 45
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 30
            ws.column_dimensions['E'].width = 42
            ws.column_dimensions['F'].width = 39
            ws.column_dimensions['G'].width = 39
            ws.column_dimensions['H'].width = 39

            # print('设置行高')
            ws.row_dimensions[1].height = 28
            ws.row_dimensions[2].height = 38
            ws.row_dimensions[3].height = 28
            ws.row_dimensions[4].height = 20

            # print('文本居中')
            ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
            ws['A2'].alignment = Alignment(horizontal='right', vertical='center')
            ws['B2'].alignment = Alignment(horizontal='left', vertical='center')
            ws['C3'].alignment = Alignment(horizontal='center', vertical='center')
            ws['A3'].alignment = Alignment(horizontal='center', vertical='center')
            ws['B3'].alignment = Alignment(horizontal='center', vertical='center')
            ws['A4'].alignment = Alignment(horizontal='center', vertical='center')
            objs = models.zhugedanao_zhongdianci_jiankong_taskDetail.objects.filter(
                tid_id=o_id
            )
            row = 5
            for obj in objs:
                tid = obj.id
                search = str(obj.search_engine)
                yinqing = ''
                if search == '1':
                    yinqing = '百度'
                elif search == '4':
                    yinqing = '手机百度'
                elif search == '3':
                    yinqing = '360'
                elif search == '6':
                    yinqing = '手机360'
                objs_detail = models.zhugedanao_zhongdianci_jiankong_taskDetailData.objects.filter(
                    tid__tid__user_id=user_id,
                    tid=tid
                ).order_by('create_time')[0:3]
                if objs_detail:
                    column_p = 4
                    for obj_data in objs_detail:
                        create_time = obj_data.create_time
                        paiming = obj_data.paiming
                        ws.cell(row=3, column=column_p, value="{create_time}".format(create_time=create_time))
                        ws.cell(row=row, column=column_p, value="{paiming}".format(paiming=paiming))
                        column_p += 1
                ws.cell(row=2, column=2, value="任务名")
                ws.cell(row=row, column=1, value="{keywords}".format(keywords=obj.keyword))
                ws.cell(row=row, column=3, value="{search_engine}".format(search_engine=yinqing))
                ws.cell(row=row, column=2, value="{lianjie}".format(lianjie=obj.lianjie))
                row += 1
            task_name = objs[0].tid.task_name
            shijainchuo = int(time.time())
            excel_name = task_name + '_' + str(shijainchuo)
            wb.save(os.path.join(os.getcwd(), 'statics', 'zhugedanao', 'zhongDianCiExcel', '{}.xlsx'.format(excel_name)))
            response.data = {'excel_name': 'http://api.zhugeyingxiao.com/' + os.path.join('statics', 'zhugedanao', 'zhongDianCiExcel','{}.xlsx'.format(excel_name))}
            response.msg = '生成成功'
            response.code = 200
    else:
        response.code = 402
        response.msg = "请求异常"
    return JsonResponse(response.__dict__)

